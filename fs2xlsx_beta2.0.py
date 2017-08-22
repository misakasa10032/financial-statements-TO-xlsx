# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import os
import re
from openpyxl import Workbook
import requests
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver

num2char = {'0':'零', '1':'一', '2':'二', '3':'三', '4':'四', '5':'五', '6':'六', '7':'七', '8':'八', '9':'九'} #dedicated to huahua
code = input('请输入股票代码(例如，3908)：')
year = input('请输入查询年份(例如，2014)：')
cover = input('请输入报表类型(年报/中报)：')
year_char = []
for k in range(len(year)):
	char = num2char[year[k]]
	year_char.append(char)
year_cnh = ''.join(year_char)

print('**********************Phase I : 正在查询报表中，请稍候**********************')
if cover == '年报':
	proposed_scope = ['全年業績', '年度業績', '全年之業績', '年度之業績']
else:
	if cover == '中报':
		proposed_scope = ['中期業績', '中期之業績']
	else:
		cover = input('您输入的报表类型无效，请重新输入(示例：年报)：')

url_1 = 'http://www.hkexnews.hk/listedco/listconews/advancedsearch/search_active_main_c.aspx'
driver = webdriver.Chrome()
driver.get(url_1)
driver.find_element_by_xpath("//input[@name = 'ctl00$txt_stock_code']").send_keys(code)
id = 0
flag = 0
for scope in proposed_scope:
	if flag == 0:
		driver.find_element_by_xpath("//input[@name = 'ctl00$txtKeyWord']").send_keys(scope)
		driver.find_element_by_xpath("//img[@src = '/image/search_c.gif']/parent::a").click()
		String = driver.page_source
		soup_0 = BeautifulSoup(String, 'lxml')
		sml_0 = soup_0.find_all('span', attrs = {'id':'ctl00_gvMain_ctl01_lblNoMatch'})
		if len(sml_0) == 0:
			proposed_text = soup_0.findAll(name = 'a', attrs = {'id': re.compile(r'\w+hlTitle')})
			flag_1 = 0
			for term in range(len(proposed_text)):
				if year_cnh in proposed_text[term].string:
					passway = proposed_text[term].string
					id = 1
					flag = 1
					flag_1 = 1
					break
				else:
					continue
			if flag_1 == 0:
				driver.get(url_1)
				driver.find_element_by_xpath("//input[@name = 'ctl00$txt_stock_code']").send_keys(code)
				continue
		else:
			driver.get(url_1)
			driver.find_element_by_xpath("//input[@name = 'ctl00$txt_stock_code']").send_keys(code)
			continue
	else:
		break
if id == 0:
	print('未能找到符合要求的报表')
	exit()
else:
	branch = soup_0.find_all(name = 'a', text = passway)
	url_2 = 'http://www.hkexnews.hk' + branch[0]['href']
	r = requests.get(url_2)
	with open(code +'.pdf', 'wb') as file:
		file.write(r.content)

print('**********************Phase II : 正在将报表转为html中，请稍候**********************')
url_3 = 'http://www.pdfdo.com/pdf-to-html.aspx'
driver.get(url_3)
driver.find_element_by_xpath("//input[@type = 'file']").send_keys('C:/Users/3.5/Desktop/' + code + '.pdf')
driver.find_element_by_xpath("//input[@type = 'submit']").click()
WebDriverWait(driver,300).until(lambda x: x.find_element_by_id("divFile").is_displayed())
branch = driver.find_element_by_xpath("//span[@id = 'ctl00_content_lblNew']/a")
url_2 = branch.get_attribute('href')
r = requests.get(url_2)
with open(code + '.html', 'wb') as file:
	file.write(r.content)
driver.quit()

print('**********************Phase III : 正在将报表转为xlsx中，请稍候**********************')
file = open('C:/Users/3.5/Desktop/' + code +'.html', 'r', encoding = 'utf-8')
str_1 = file.read()
soup = BeautifulSoup(str_1,'lxml')
soup_list = soup.find_all(name = 'div', attrs = {'class': 'view'})
class_name = []
for branch in soup_list:
	class_name.append(tuple(branch.parent['class']))
class_name = sorted(set(class_name), key = class_name.index)
sml = []
for name in class_name:
	sml_0 = soup.find_all('div', attrs = {'class': name})
	sml.extend(sml_0)
count = 0
w = Workbook()
sml.pop()
for term in sml:
	existence = 0
	exp_temp = term.find_all('span')
	loc_temp = []
	left_temp = []
	type_temp = exp_temp[-1]['class']
	std_list_0 = range(len(exp_temp))
	std_list_2 = list(reversed(std_list_0))
	pop_list_0 = []
	for num in std_list_2:
		if exp_temp[num]['class'] == type_temp:
			pop_list_0.append(num)
		else:
			break
	for num in pop_list_0:
		del exp_temp[num]
	for item in exp_temp:
		if 'style' in list(item.parent.attrs.keys()):
			loc_1temp1 = item.parent['style']
			loc_temp.append(loc_1temp1)
		else:
			loc_1temp2 = item.parent.parent['style']
			loc_temp.append(loc_1temp2)
	for item in loc_temp:
		style_1temp = re.findall('\d+\.*\d*' + 'em', item)
		flo_lefttemp = float(style_1temp[1].replace('em', ''))
		left_temp.append(flo_lefttemp)
	for j in range(len(left_temp) - 1):
		gap_temp = left_temp[j + 1] - left_temp[j]
		if gap_temp > 2.5:
			existence = 1
	if existence == 1:
		count += 1
		exp = term.find_all('span')
		ich = 0
		type_0 = exp[-1]['class']
		std_list_1 = range(len(exp))
		std_list = list(reversed(std_list_1))
		pop_list = []
		for num in std_list:
			if exp[num]['class'] == type_0:
				pop_list.append(num)
			else:
				break
		for num in pop_list:
			del exp[num]
		loc = []
		string = []
		loc_tup = []
		for item in exp:
			string_1 = item.string
			string_1 = string_1.replace('\xa0','')
			string_1 = string_1.strip()
			if re.sub('已發行：', '', string_1) == '':
				ich = 1
			string.append(string_1)
			if 'style' in list(item.parent.attrs.keys()):
				loc_1 = item.parent['style']
				loc.append(loc_1)
			else:
				loc_tm = item.parent.parent['style']
				loc.append(loc_tm)
		if sml.index(term) == 0:
			for m in range(len(string)):
				if string[m] == '1':
					string[m] = ''
		top_1 =[]
		left_1 = []
		for item in loc:
			style_1 = re.findall('\d+\.*\d*' + 'em', item)
			flo_top = float(style_1[0].replace('em', ''))
			flo_left = float(style_1[1].replace('em', ''))
			top_1.append(flo_top)
			left_1.append(flo_left)
			loc_tup.append((flo_top,flo_left))
		top = list(set(top_1))
		left = list(set(left_1))
		top_n = sorted(top) 
		left_n = sorted(left)
		top_occ = []
		left_occ = []
		top_tim = 0
		left_tim = 0
		for j in range(len(top_n) - 1):
			gap = top_n[j + 1] - top_n[j]
			if gap > 0.5:
				top_occ.append(j)
				top_tim += 1
		for j in range(len(left_n) - 1):
			gap = left_n[j + 1] - left_n[j]
			if gap > 2.5:
				left_occ.append(j)
				left_tim += 1
		top_occ.insert(0, -1)
		left_occ.insert(0, -1)
		top_occ.append(len(top_n) - 1)
		left_occ.append(len(left_n) - 1)
		top_std = range(len(top_occ) - 1)
		top_dict = {}
		left_std = range(len(left_occ) - 1)
		left_dict = {}
		for j in top_std:
			top_dict[j] = []
			if top_occ[j] + 1 == top_occ[j + 1]:
				top_dict[j].append(top_n[top_occ[j + 1]])
			else:
				for i in range(top_occ[j] + 1, top_occ[j + 1] + 1):
					top_dict[j].append(top_n[i])
			top_dict[j] = tuple(top_dict[j])
		for j in left_std:
			left_dict[j] = []
			if left_occ[j] + 1 == left_occ[j + 1]:
				left_dict[j].append(left_n[left_occ[j + 1]])
			else:
				for i in range(left_occ[j] + 1, left_occ[j + 1] + 1):
					left_dict[j].append(left_n[i])
			left_dict[j] = tuple(left_dict[j])
		top_inverse = {v:k for k,v in top_dict.items()}
		left_inverse = {v:k for k,v in left_dict.items()}
		loc_2 = []
		for item_1 in loc_tup:
			for item_2 in list(top_dict.values()):
				if item_1[0] in item_2:
					row = top_inverse[item_2]
					row_indicator = 1
			for item_3 in list(left_dict.values()):
				if item_1[1] in item_3:
					column = left_inverse[item_3]
					column_indicator = 1
			loc_2.append((row,column))
		loc_3 = sorted(set(loc_2), key = loc_2.index)
		ent = []
		for pos in loc_2:
			lol = []
			for pos_0 in enumerate(loc_2):
				if pos_0[1] == pos:
					lol.append(pos_0[0])
			ent.append(lol)
		ent_2 = []
		for k in ent:
			k_1 = tuple(k)
			ent_2.append(k_1)
		ent_3 = tuple(ent_2)
		ent_1 = sorted(set(ent_3), key = ent_3.index)
		String_2 = []
		for item in ent_1:
			if len(item) == 1:
				String_2.append(string[item[0]])
			if len(item) > 1:
				var = []
				var_2 = []
				left_dict = {}
				for a in item:
					left_dict[a] = left_1[a]
				var_1 = sorted(left_dict.items(), key = lambda item: item[1])
				for term in var_1:
					var_2.append(term[0])
				var_id = 0
				for x in var_2:
					var.append(string[x] + '?')
				String_2.append(''.join(var))
		Name = 'Sheet' + str(count)
		sheet = w.create_sheet()
		for i in range(len(String_2)):
			loc = loc_3[i]
			goal = String_2[i]
			goal_0 = re.split(' ', goal)
			identify = 0
			if len(goal_0) == 1:
				identify = 1
			else:
				for indexs in range(len(goal_0) - 1):
					if goal_0[indexs] != goal_0[indexs + 1]:
						identify = 1
			if identify == 0:
				del String_2[i]
				del loc_3[i]
				for v in range(len(goal_0)):
					String_2.insert(i + v, goal_0[v])
					loc_3.insert(i + v, (loc[0], loc[1] + v + 1))
		for i in range(len(String_2)):
			loc = loc_3[i]
			sheet.cell(row = loc[0] + 1, column = loc[1] + 1).value = String_2[i]
w.remove_sheet(w.get_sheet_by_name(w.get_sheet_names()[0]))

print('**********************Phase IV : 正在优化报表中，请稍候**********************')
sheet_len = len(w.get_sheet_names())
outliers_loc = []
outliers_val = []
outliers_list = []
sheet_names = w.get_sheet_names()	
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	iden = 0
	iden_list = []
	Null_list = []
	col_data = list(sheet.columns)[0]
	for k in range(len(col_data)):
		if col_data[k].value is None:
			iden = 1
			Null_list.append(k)
	iden_list.append(iden)
	gap1_list = []
	for j in range(len(Null_list) - 1):
		gap_1 = Null_list[j + 1] - Null_list[j]
		gap1_list.append(gap_1)
	temp = []
	initial = 0
	for d in range(len(gap1_list)):
		if gap1_list[d] == 1:
			if d == len(gap1_list) - 1:
				initial += 1
				temp.append(initial)
			else:
				if gap1_list[d + 1] != 1:
					initial += 1
					temp.append(initial)
				else:
					initial += 1
		else:
			initial = 0
			temp.append(initial)
	cache = []
	for k in range(len(temp)):
		if temp[k] != 0:
			cache.append(temp[k] + 1)
	def func(list, s):
		if list[s] == 0:
			return 1
		else:
			return list[s]
	sup = []
	for s in range(len(temp)):
		if s == 0:
			if temp[s] != 0:
				sup.append(Null_list[0])
		else:
			if temp[s] != 0:
				sum = 0
				for h in range(s):
					sum += func(temp, h)
				sup.append(Null_list[sum])
	edge_list = []
	value_list = []
	loc_list = []
	for x in range(len(sup)):
		edge_1 = sup[x] + cache[x] - 1
		edge_list.append(edge_1)
		edge_2 = sup[x] + cache[x] - 2
		edge_list.append(edge_2)
		if sup[x] + cache[x] - 3 > - 1:
			edge_3 = sup[x] + cache[x] - 3
			edge_list.append(edge_3)
	ncols = sheet.max_column
	dich = []
	for row_0 in edge_list:
		for col_0 in range(ncols):
			cellv = sheet.cell(row = row_0 + 1, column = col_0 + 1).value
			if not cellv is None: 
				if (re.search('\d+', cellv) is None and not re.search(' ', cellv) is None and i != 0):
					deutsch = (cellv, (row_0, col_0))
					dich.append(deutsch)
	dich = sorted(set(dich), key = dich.index)
	for ich in dich:
		value_list.append(ich[0])
		loc_list.append(ich[1])
	outliers_val.append(value_list)
	outliers_loc.append(loc_list)
	if value_list != []:
		outliers_list.append(i)
for i in outliers_list:
	sheet = w.get_sheet_by_name(sheet_names[i])
	ncols = sheet.max_column
	for term in range(len(outliers_loc[i])):
		if ' ' in outliers_val[i][term]:
			move = []
			ha_0 = re.split(' ', outliers_val[i][term])
			if outliers_loc[i][term][1] + 1 == ncols:
				move.append(ncols)
			for col in range(outliers_loc[i][term][1] + 2, ncols + 1):
				if (not sheet.cell(row = outliers_loc[i][term][0] + 1, column = col).value is None):
					move.append(col)
			if move == []:
				for danke in range(len(ha_0)):
					sheet.cell(row = outliers_loc[i][term][0] + 1, column = outliers_loc[i][term][1] + danke + 1).value = ha_0[danke]
			else:
				for col in reversed(range(len(move))):
					word = sheet.cell(row = outliers_loc[i][term][0] + 1, column = move[col]).value
					if not word is None:
						sheet.cell(row = outliers_loc[i][term][0] + 1, column = move[col]).value = ''
						sheet.cell(row = outliers_loc[i][term][0] + 1, column = outliers_loc[i][term][1] + len(ha_0) + col + 1).value = word
				for ink in range(len(ha_0)):
					sheet.cell(row = outliers_loc[i][term][0] + 1, column = outliers_loc[i][term][1] + ink + 1).value =  ha_0[ink]
for i in range(len(iden_list)):
	if iden_list[i] == 0:
		w.remove_sheet(w.get_sheet_by_name(w.get_sheet_names()[i]))
outliers_val = []
outliers_loc = []
outliers_list = []
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	iden = 0
	Null_list = []
	col_data = list(sheet.columns)[0]
	for k in range(len(col_data)):
		if col_data[k].value is None:
			iden = 1
			Null_list.append(k)		
	edge_list = []
	value_list = []
	loc_list = []
	gap1_list = []
	for j in range(len(Null_list) - 1):
		gap_1 = Null_list[j + 1] - Null_list[j]
		gap1_list.append(gap_1)
	temp = []
	initial = 0
	for d in range(len(gap1_list)):
		if gap1_list[d] == 1:
			if d == len(gap1_list) - 1:
				initial += 1
				temp.append(initial)
			else:
				if gap1_list[d + 1] != 1:
					initial += 1
					temp.append(initial)
				else:
					initial += 1
		else:
			initial = 0
			temp.append(initial)
	cache = []
	for k in range(len(temp)):
		if temp[k] != 0:
			cache.append(temp[k] + 1)
	def func(list, s):
		if list[s] == 0:
			return 1
		else:
			return list[s]
	sup = []
	for s in range(len(temp)):
		if s == 0:
			if temp[s] != 0:
				sup.append(Null_list[0])
		else:
			if temp[s] != 0:
				sum = 0
				for h in range(s):
					sum += func(temp, h)
				sup.append(Null_list[sum])
	for x in range(len(sup)):
		edge_1 = sup[x] + cache[x] - 1
		edge_list.append(edge_1)
		edge_2 = sup[x] + cache[x] - 2
		edge_list.append(edge_2)
		if sup[x] + cache[x] - 3 > - 1:
			edge_3 = sup[x] + cache[x] - 3
			edge_list.append(edge_3)
	ncols = sheet.max_column
	dich = []
	for row_0 in edge_list:
		for col_0 in range(ncols):
			cellv = sheet.cell(row = row_0 + 1, column = col_0 + 1).value
			if not cellv is None: 
				if ((re.search('\d+', cellv) is None and len(re.findall('[\u4e00-\u9fa5]', cellv)) != 0) or ('年' in cellv and not '年度' in cellv and not '全年' in cellv)):
					if '?' in cellv.rstrip('?'):
						deutsch = (cellv, (row_0, col_0))
						dich.append(deutsch)
	dich = sorted(set(dich), key = dich.index)
	for ich in dich:
		value_list.append(ich[0])
		loc_list.append(ich[1])
	outliers_val.append(value_list)
	outliers_loc.append(loc_list)
	if value_list != []:
		outliers_list.append(i)
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	ncols = sheet.max_column
	for term in range(len(outliers_loc[i])):
		year_list = []
		year_bank = []
		ha_0 = re.split('\?', outliers_val[i][term])
		if '' in ha_0:
			ha_0.remove('')
		if '年' in ha_0:
			for item in range(len(ha_0)):
				if ha_0[item] == '年':
					year_list.append(item)
			year_list.insert(0, -1)
			for k in range(len(year_list) - 1):
				if year_list[k] + 1 == year_list[k + 1] - 1:
					year_bank.append(ha_0[year_list[k] + 1] + '年')
				if year_list[k] + 1 < year_list[k + 1] - 1:
					small_piece = []
					for j in range(year_list[k] + 1, year_list[k + 1]):
						small_piece.append(ha_0[j])
					year_bank.append(''.join(small_piece) + '年')
			ha_0 = year_bank
		idens = 0
		alter = ['(', ')']
		for ite in alter:
			if ite in ha_0:
				idens = 1		
		if ('?' in outliers_val[i][term] and idens == 0):
			move = []
			sas = []
			for col in range(outliers_loc[i][term][1] + 2, ncols + 1):
				if (not sheet.cell(row = outliers_loc[i][term][0] + 1, column = col).value is None or sheet.cell(row = outliers_loc[i][term][0] + 1, column = col).value != ''):
					move.append(col)
				if ((outliers_loc[i][term][0], col - 1) in outliers_loc[i] and col - outliers_loc[i][term][1] - 1 < len(ha_0)):
					sas.append(col)
			if move == []:
				for danke in range(len(ha_0)):
					sheet.cell(row = outliers_loc[i][term][0] + 1, column = outliers_loc[i][term][1] + danke + 1).value = ha_0[danke]
			else:
				for col in reversed(range(len(move))):
					word = sheet.cell(row = outliers_loc[i][term][0] + 1, column = move[col]).value
					if not word is None:
						sheet.cell(row = outliers_loc[i][term][0] + 1, column = move[col]).value = ''
						sheet.cell(row = outliers_loc[i][term][0] + 1, column = outliers_loc[i][term][1] + len(ha_0) + col + 1).value = word
					if move[col] in sas:
						outliers_loc[i][term + sas.index(move[col]) + 1] = (outliers_loc[i][term][0], outliers_loc[i][term][1] + len(ha_0) + col)
				for ink in range(len(ha_0)):
					sheet.cell(row = outliers_loc[i][term][0] + 1, column = outliers_loc[i][term][1] + ink + 1).value =  ha_0[ink]
outvalue_loc = []
outvalue = []
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	nrows = sheet.max_row
	ncols = sheet.max_column
	getlist_1 = []
	getlist_2 = []
	for row_0 in range(1, nrows + 1):
		for column_0 in range(1, ncols + 1):
			word = sheet.cell(row = row_0, column = column_0).value
			if (not word is None and '?' in word):
				ha_1 = re.split('\?', word)
				ha_1.pop()
				idens = 0
				alter = ['(', ')', '1', '2', '3', '4', '5', '6', '7', '8', '9']
				for ite in alter:
					if ite in ha_1:
						idens = 1
				if ' ' in word:
					word= word.replace(' ','?')
				if (re.sub('\.', '', re.sub('%', '', re.sub('－', '', re.sub('–', '', re.sub('-', '', re.sub('\?', '', re.sub('—', '', (re.sub('\)', '', re.sub('\(', '', re.sub(',','', re.sub('\d+','', word)))))))))))) == '' and idens == 0):
						getlist_1.append((row_0, column_0))
						getlist_2.append(word)
	outvalue_loc.append(getlist_1)
	outvalue.append(getlist_2)
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	ncols = sheet.max_column
	for term in range(len(outvalue_loc[i])):
		if '?' in outvalue[i][term]:
			move = []
			sas = []
			ha_0 = re.split('\?', outvalue[i][term])
			if '' in ha_0:
				ha_0.remove('')
			for col in range(outvalue_loc[i][term][1] + 1, ncols + 1):
				if (not sheet.cell(row = outvalue_loc[i][term][0], column = col).value is None or sheet.cell(row = outvalue_loc[i][term][0], column = col).value != ''):
					move.append(col)
				if ((outvalue_loc[i][term][0], col) in outvalue_loc[i] and col - outvalue_loc[i][term][1] < len(ha_0)):
					sas.append(col)
			if move == []:
				for danke in range(len(ha_0)):
					sheet.cell(row = outvalue_loc[i][term][0], column = outvalue_loc[i][term][1] + danke).value = ha_0[danke]
			else:
				for col in reversed(range(len(move))):
					word = sheet.cell(row = outvalue_loc[i][term][0], column = move[col]).value
					if not word is None:
						sheet.cell(row = outvalue_loc[i][term][0], column = move[col]).value = ''
						sheet.cell(row = outvalue_loc[i][term][0], column = outvalue_loc[i][term][1] + len(ha_0) + col).value = word
					if move[col] in sas:
						outvalue_loc[i][term + sas.index(move[col]) + 1] = (outvalue_loc[i][term][0], outvalue_loc[i][term][1] + len(ha_0) + col)
				for ink in range(len(ha_0)):
					sheet.cell(row = outvalue_loc[i][term][0], column = outvalue_loc[i][term][1] + ink).value =  ha_0[ink]
outvalue_loc = []
outvalue = []
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	nrows = sheet.max_row
	ncols = sheet.max_column
	getlist_1 = []
	getlist_2 = []
	for row_0 in range(1, nrows + 1):
		for column_0 in range(1, ncols + 1):
			word = sheet.cell(row = row_0, column = column_0).value
			if (not word is None and ' ' in word):
				ha_1 = re.split(' ', word)
				idens = 0
				alter = ['(', ')', '1', '2', '3', '4', '5', '6', '7', '8', '9']
				for ite in alter:
					if ite in ha_1:
						idens = 1
				if (re.sub('\.', '', re.sub('%', '', re.sub('－', '', re.sub('–', '', re.sub('-', '', re.sub(' ', '', re.sub('—', '', (re.sub('\)', '', re.sub('\(', '', re.sub(',','', re.sub('\d+','', word)))))))))))) == '' and idens == 0):
						getlist_1.append((row_0, column_0))
						getlist_2.append(word)
	outvalue_loc.append(getlist_1)
	outvalue.append(getlist_2)
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	ncols = sheet.max_column
	for term in range(len(outvalue_loc[i])):
		if ' ' in outvalue[i][term]:
			move = []
			sas = []
			ha_0 = re.split(' ', outvalue[i][term])
			if '' in ha_0:
				ha_0.remove('')
			for col in range(outvalue_loc[i][term][1] + 1, ncols + 1):
				if (not sheet.cell(row = outvalue_loc[i][term][0], column = col).value is None or sheet.cell(row = outvalue_loc[i][term][0], column = col).value != ''):
					move.append(col)
				if ((outvalue_loc[i][term][0], col) in outvalue_loc[i] and col - outvalue_loc[i][term][1] < len(ha_0)):
					sas.append(col)
			if move == []:
				for danke in range(len(ha_0)):
					sheet.cell(row = outvalue_loc[i][term][0], column = outvalue_loc[i][term][1] + danke).value = ha_0[danke]
			else:
				for col in reversed(range(len(move))):
					word = sheet.cell(row = outvalue_loc[i][term][0], column = move[col]).value
					if not word is None:
						sheet.cell(row = outvalue_loc[i][term][0], column = move[col]).value = ''
						sheet.cell(row = outvalue_loc[i][term][0], column = outvalue_loc[i][term][1] + len(ha_0) + col).value = word
					if move[col] in sas:
						outvalue_loc[i][term + sas.index(move[col]) + 1] = (outvalue_loc[i][term][0], outvalue_loc[i][term][1] + len(ha_0) + col)
				for ink in range(len(ha_0)):
					sheet.cell(row = outvalue_loc[i][term][0], column = outvalue_loc[i][term][1] + ink).value =  ha_0[ink]
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	ncols = sheet.max_column
	nrows = sheet.max_row
	for row_0 in range(1, nrows + 1):
		for col_0 in range(1, ncols + 1):
			cellval = sheet.cell(row = row_0, column = col_0).value
			if (cellval == '（' and not sheet.cell(row = row_0, column = col_0 + 1).value is None):
				sheet.cell(row = row_0, column = col_0).value = ''
				sheet.cell(row = row_0, column = col_0 + 1).value = '（' + sheet.cell(row = row_0, column = col_0 + 1).value
			if (cellval == '）' and not sheet.cell(row = row_0, column = col_0 - 1).value is None):
				sheet.cell(row = row_0, column = col_0).value = ''
				sheet.cell(row = row_0, column = col_0 - 1).value = sheet.cell(row = row_0, column = col_0 - 1).value + '）'
str_loc = []
str_list = []
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	ncols = sheet.max_column
	final = list(sheet.columns)[ncols - 1]
	spc = []
	spc_loc = []
	for term in range(1, len(final) + 1):
		word = final[term - 1].value
		if (not word is None and word != ''):
			spc.append(word)
			spc_loc.append(term)
	if (len(spc) == 2 or len(spc) == 1):
		whelm = list(sheet.rows)[spc_loc[0] - 1]
		for term in range(2, len(whelm) + 1):
			word = whelm[term - 1].value
			if term == 2:
				if (not word is None and word != ''):
					for k in range(len(spc)):
						if (not sheet.cell(row = spc_loc[k], column = 1).value is None and not sheet.cell(row = spc_loc[k], column = 2).value is None):
							sheet.cell(row = spc_loc[k], column = 1).value = sheet.cell(row = spc_loc[k], column = 1).value + sheet.cell(row = spc_loc[k], column = 2).value
			if term > 2:
				if (not word is None and word != ''):
					for k in range(len(spc)):
						sheet.cell(row = spc_loc[k], column = term).value = ''
						sheet.cell(row = spc_loc[k], column = term - 1).value = word
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	nrows = sheet.max_row
	ncols = sheet.max_column
	for k in range(nrows):
		bazinga = []
		ful = list(sheet.rows)[k]
		ids = 0
		dd = 0
		rank = []
		break_point = 0
		for term in ful:
			if (not term.value is None and term.value != ''):
				bazinga.append(term)
				rank.append(ful.index(term))
		for h in range(1, len(rank)):
			gap = rank[h] - rank[h - 1]
			if gap != 1:
				dd = 1
				break_point = rank[h-1]
		if re.sub('\.', '', re.sub('%', '', re.sub('－', '', re.sub('–', '', re.sub('-', '', re.sub('—', '', (re.sub('\)', '', re.sub('\(', '', re.sub(',','', re.sub('\d+','', bazinga[-1].value))))))))))) == '': 
			ids = 1
		if (ids == 1 and dd == 0):
			if (len(bazinga) > 2 and len(bazinga) < ncols - 1 and ful[-1].value is None):
				word = bazinga[-1].value
				sheet.cell(row = k + 1, column = ncols).value = word
				sheet.cell(row = k + 1, column = ful.index(bazinga[-1]) + 1).value = ''
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	nrows = sheet.max_row
	ncols = sheet.max_column
	ful = list(sheet.columns)[ncols - 1]
	bazinga = []
	for term in ful:
		if (not term.value is None and term.value != ''):
			bazinga.append(term)
	idd = 0
	for k in range(len(bazinga)):
		if re.sub('\.', '', re.sub('%', '', re.sub('－', '', re.sub('–', '', re.sub('-', '', re.sub('—', '', (re.sub('\)', '', re.sub('\(', '', re.sub(',','', re.sub('\d+','', bazinga[k].value))))))))))) != '': 
			idd = 1
	if idd == 0:
		for liebe in bazinga:
			liebe_list = list(sheet.rows)[ful.index(liebe)]
			real = []
			for fake in liebe_list:
				if (not fake.value is None and fake.value != ''):
					real.append(fake)
			goal_col = liebe_list.index(real[-1])
			val = liebe.value
			liebe.value = ''
			sheet.cell(row = ful.index(liebe) + 1, column = goal_col).value = val	
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	nrows = sheet.max_row
	ncols = sheet.max_column
	for k in range(nrows):
		bazinga = []
		ful = list(sheet.rows)[k]
		ids = 0
		dd = 0
		rank = []
		break_point = 0
		for term in ful:
			if (not term.value is None and term.value != ''):
				bazinga.append(term)
				rank.append(ful.index(term))
		for h in range(1, len(rank)):
			gap = rank[h] - rank[h - 1]
			if gap != 1:
				dd = 1
				break_point = rank[h-1]
			if gap == 2:
				for t in reversed(range(1, break_point + 1)):
					word = ful[t].value
					ful[t].value = ''
					sheet.cell(row = k + 1, column = t + 2).value = word
for i in range(sheet_len):
	sheet = w.get_sheet_by_name(sheet_names[i])
	nrows = sheet.max_row
	ncols = sheet.max_column
	for row_0 in range(1, nrows + 1):
		for col_0 in range(1, ncols + 1):
			word = sheet.cell(row = row_0, column = col_0).value
			if not word is None:
				word = re.sub('\?', '', word)
				sheet.cell(row = row_0, column = col_0).value = word
w.save(code + '.xlsx')
print('**********************Congratulations : 报表转化完成！**********************')	# I love huahua!




	
	
	
	
	
			
		


	

	
